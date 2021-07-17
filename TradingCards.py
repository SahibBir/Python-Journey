"""
Submitted By: SAHIB BIR SINGH BHATIA
Student ID: 201547831
COMP 517 - Assignment 5 - 2020-21 - JAN21 - CA Assignment 5 - TRADING CARDS
"""

# Importing openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


# Card Class
class Card(object):
    """
    Card Class contains three methods namely:
    1. __init__(self, theName, theType,theHP,theMoves,isShiny) which initialises the Class instance
    2. __getitem__(self, item) which helps in return the created class instance as a dictionary
    3. __str__(self) which prints the card's information
    IMP: Please note that theMoves is a list of tuples like [('move 1 name','damage: value of damage'),...and so on]
    """

    typeList = ["Magi", "Water", "Fire", "Earth", "Air", "Astral"]  # List containing valid card types

    def __init__(self, theName: str, theType: str, theHP: int, theMoves: list, isShiny: int,):
        """__init__(self, theName, theType,theHP,theMoves,isShiny) takes the input as follows:
        1.theName as string
        2.theType as string
        3.theHp as integer
        4.theMoves as a list of tuples like [('move 1 name','damage: value of damage'),...and so on]"""
        movesStatusflag = False  # Flag represents whether the list passed as moves is valid or not.[Initially False]
        # checking if input given by user while creating a card object is valid or not
        if theType in Card.typeList and \
                theHP >= 0 and \
                isShiny in [0, 1] and \
                len(theMoves) != 0 and \
                type(theMoves) == list:  # if everything is valid
            for p in range(len(theMoves)):
                if len(theMoves[p]) != 0:  # checking if the tuples inside in the moves list are valid or not
                    if str(theMoves[p][1]).split(":")[-1] != "":  # checking if damage is given or not
                        if str(theMoves[p][0]).split(",")[0] != "":  # checking if move name  is given or not
                            movesStatusflag = True  # if move name and damage is present then update flag
                        else:
                            movesStatusflag = False  # if moves name  is invalid then update flag
                    else:
                        movesStatusflag = False  # if damage is invalid then update flag
                else:
                    movesStatusflag = False  # if length moves list is invalid that is of no moves are given by the user
            if movesStatusflag:  # if flag is true create a dictionary
                self.ca = {"Name": theName, "Type": theType, "HP": theHP, "Shiny": isShiny, "Moves": theMoves}
            else:  # print given moves/damage is invalid
                print("Invalid Moves or Damage.")
        elif theHP < 0:  # if given HP is invalid
            print("Health Point cannot be less than zero.")
        elif isShiny not in [0, 1]:  # if given Shiny value is invalid
            print("Shiny attribute can be either 0 or 1.")
        elif type(theMoves) != list:  # if moves is not given in a list of tuples
            print("Please provide moves in a list in the format as [('move1',damage1:value of damage'),...and so on]")
        else:  # if given card type not present in the typeList mentioned above
            print("Please select a valid card type from one of the below mentioned type: \n""Magi"
                  "\nWater\nFire\nEarth\nAir\nAstral\n")

    def __getitem__(self, item):
        """# Returns the class object created by user as a dictionary"""
        return self.ca[item]

    def __str__(self):
        """__str__(self): prints the card's information in a sensible manner"""
        try:  # try except statements to catch and print errors and corner cases
            if self.ca["Shiny"] == 1:  # checking if the card is shiny and printing accordingly
                return "Card has been created.Please find card details as follows:\nCard Name=> {} \nType=> {} \nHP=> "\
                       "{} \nShiny=> {} \nMoves=> {} \n---------------------------------".format(self.ca["Name"],
                                                                                                 self.ca["Type"],
                                                                                                 self.ca["HP"], True,
                                                                                                 self.ca["Moves"])
            else:  # checking if the card is not shiny and printing accordingly
                return "Card has been created.Please find card details as follows:\nCard Name=> {} \nType=> {} \nHP=> "\
                       "{} \nShiny=> {} \nMoves=> {} \n---------------------------------".format(self.ca["Name"],
                                                                                                 self.ca["Type"],
                                                                                                 self.ca["HP"], False,
                                                                                                 self.ca["Moves"])
        except ValueError:
            return "Please try again\n"


# Deck Class
class Deck(object):
    """Deck Class contains the following methods:
    1.__init__(self) which initialise the Deck instance
    2.inputFromFile(self, fileName) which takes an excel file containing cards as input and adds these card to deck
    3.__str__(self) provides the total number of cards, the total number of shiny cards, and the average Damage value of
     the entire deck.
    4.addCard(self, theCard) adds the requested Card to the deck
    5.rmCard(self, theCard) removes the requested Card from the deck
    6.getMostPowerful(self) returns the most powerful card in the deck
    7.getAverageDamage(self) returns the average damage of the deck
    8.viewAllCards(self) prints the information of all cards in a tabular format
    9.viewAllShinyCards(self) prints the information of all shiny cards
    10.viewAllByType(self, theType) prints the information of all cards of the requested type
    11.getCards(self) returns all cards held in the deck as a collecton of type of dictionary
    12.saveToFile(self, fileName) saves the deck to the requested file
    """

    def __init__(self):
        """__init__(self) initialises an empty dictionary,empty list and an active sheet"""
        self.cc = {}  # Initialising any empty dictionary
        self.keysList = []  # Initialising an empty list which will contains the keys of dict as mentioned in the name
        self.activeSheet = None  # Initialising an empty sheet

    # Func inputFromFile takes an excel file as input
    def inputFromFile(self, fileName):
        """Function to populate the empty initialised Deck with the cards that are present in
        the .xlsx file <fileName>."""
        try:  # File Found
            wb = load_workbook(filename="/Users/sahibbhatia/Desktop/Programming_Fundamentals/{}".format(fileName))
            """Please change the path above accordingly"""
            self.activeSheet = wb.active
            # Iterating through the active sheet of the input file
            for row in self.activeSheet.iter_rows(min_row=2,
                                                  max_row=wb.active.max_row,
                                                  min_col=1,
                                                  max_col=14,
                                                  values_only=True):
                cardName = row[0]
                cardProperties = {
                    "Type": row[1],
                    "HP": row[2],
                    "Shiny": row[3],
                    "Moves": [(row[4], "Damage:" + str(row[5])), (row[6], "Damage:" + str(row[7])),
                              (row[8], "Damage:" + str(row[9])), (row[10], "Damage:" + str(row[11])),
                              (row[12], "Damage:" + str(row[13]))]
                }
                for i in cardProperties["Moves"]:  # Removing empty columns
                    if i[0] is None:
                        cardProperties["Moves"].remove(i)

                self.cc[cardName] = cardProperties  # Creating a dictionary of cards with card name as key
                self.keysList = list(self.cc.keys())  # Creating a list containing keys of the dictionary
        except Exception as e:  # File not found
            print(e)
            print("Please try again.\n")

    def __str__(self):
        """__str__(self) providing string representation of the deck"""
        d = self.getAverageDamage()  # calling func getAverageDamage() to fetch average Value of deck
        d = d[-5:]
        shinyCardlist = []  # Initialising an empty to list
        for m in range(0, len(self.keysList)):
            isShiny = self.cc[self.keysList[m]]["Shiny"]
            if isShiny == 1:
                shinyCardlist.append(self.keysList[m])  # Adding any shiny card in the list
            else:
                continue

        return "{:<20} | {:<20} | {:<20}\n{:<21} | {:<27} | {:<20}\n".format('Total Number of Cards',
                                                                             'Total Number of Shiny '
                                                                             'Cards', 'Average Damage '
                                                                                      'Value of Deck',
                                                                             len(self.keysList), len(shinyCardlist), d)

    # Func addCard adds the given card to the deck
    def addCard(self, theCard):
        """addCard takes the input parameter theCard and then breaks it down to create a nested dictionary so that the
        card can be added to the deck"""
        cardName = theCard["Name"]
        cardProperties = {
            "Type": theCard["Type"],
            "HP": theCard["HP"],
            "Shiny": theCard["Shiny"],
            "Moves": theCard["Moves"]
        }
        if len(theCard["Moves"]) == 0:  # checks again if card has moves or not
            return "Your card doesn't have any moves.Please create new card."
        else:
            self.cc[cardName] = cardProperties  # adding card to the deck
            self.keysList.append(cardName)
            return "Card {} has been added to this deck".format(theCard["Name"])

    # Func rmCard removes the given card from the deck
    def rmCard(self, theCard):
        """rmCard deletes the requested card from the deck"""
        del self.cc[theCard['Name']]
        self.keysList.remove(theCard['Name'])
        return "Card:{} has been removed from deck.\nCards left are:{}".format(theCard['Name'], self.keysList)

    # Func getMostPowerful returns the most powerful card in the deck
    def getMostPowerful(self):
        """getMostPowerful calculates the most powerful card by calculating average damage values of each card in the
        deck and then comparing them with each other to return the crd which has the highest average damage value"""
        listAverageValues = []  # Initialising an empty list to keep average damage values of a card in a deck
        for i in range(0, len(self.keysList)):
            x = 0
            for j in range(0, len(self.cc[self.keysList[i]]["Moves"])):
                damage = self.cc[self.keysList[i]]["Moves"][j][1]
                damage = int(damage[7:])
                x = x + damage
            avgX = round(x / len(self.cc[self.keysList[i]]["Moves"]), 2)  # Calculating average damage of card
            listAverageValues.append(avgX)  # adding calculated average damage value to the list
        # Returning most powerful card
        return "The most powerful card in this deck is {} with average damage of {}".format(
            self.keysList[listAverageValues.index(max(listAverageValues))], max(listAverageValues))

    # Func getAverageDamage returns the average damage of the deck
    def getAverageDamage(self):
        """getAverageDamage calculates the average damage of deck by calculating the average damage of each card in the
        deck and then taking average of all those values"""
        listOfaverageValues = []  # Initialising an empty list to keep average damage values of a card in a deck
        for i in range(0, len(self.keysList)):
            x = 0
            for j in range(0, len(self.cc[self.keysList[i]]["Moves"])):
                damage = self.cc[self.keysList[i]]["Moves"][j][1]
                damage = int(damage[7:])
                x = x + damage
            avgX = round(x / len(self.cc[self.keysList[i]]["Moves"]), 2)  # Calculating average damage of card
            listOfaverageValues.append(avgX)  # adding calculated average damage value to the list
        # Calculating average damage of deck
        averageDamageofDeck = round(sum(listOfaverageValues) / (len(listOfaverageValues)), 1)
        # Returning average damage of deck
        return "The average damage of this deck is {}".format(averageDamageofDeck)

    # Func viewAllCards prints the information of all cards in a deck
    def viewAllCards(self):
        """viewAllCards prints the information of all the cards in the deck in a tabular format"""
        print("The following are the cards in this deck:")
        print("{:<12} {:<10} {:<6} {:<6} {:<30} {:<45} {:<45} {:<45} {:<45}".format('Name', 'Type', 'HP', 'Shiny',
                                                                                    'Move1:Damage1', 'Move2:Damage2',
                                                                                    'Move3:Damage3', 'Move4:Damage4',
                                                                                    'Move5:Damage5'))
        for q in range(0, len(self.keysList)):
            if len(self.cc[self.keysList[q]]["Moves"]) == 5:
                print("{:<12} {:<10} {:<6} {:<6} {:<30} {:<45} {:<45} {:<45} {:<45}".format(
                    self.keysList[q], self.cc[self.keysList[q]]["Type"], self.cc[self.keysList[q]]["HP"],
                    self.cc[self.keysList[q]]["Shiny"],
                    str(self.cc[self.keysList[q]]["Moves"][0][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][0][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][1][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][1][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][2][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][2][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][3][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][3][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][4][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][4][1]).split(":")[1]
                ))
            elif len(self.cc[self.keysList[q]]["Moves"]) == 4:
                print("{:<12} {:<10} {:<6} {:<6} {:<30} {:<45} {:<45} {:<45}".format(
                    self.keysList[q], self.cc[self.keysList[q]]["Type"], self.cc[self.keysList[q]]["HP"],
                    self.cc[self.keysList[q]]["Shiny"],
                    str(self.cc[self.keysList[q]]["Moves"][0][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][0][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][1][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][1][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][2][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][2][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][3][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][3][1]).split(":")[1]
                ))
            elif len(self.cc[self.keysList[q]]["Moves"]) == 3:
                print("{:<12} {:<10} {:<6} {:<6} {:<30} {:<45} {:<45} ".format(
                    self.keysList[q], self.cc[self.keysList[q]]["Type"], self.cc[self.keysList[q]]["HP"],
                    self.cc[self.keysList[q]]["Shiny"],
                    str(self.cc[self.keysList[q]]["Moves"][0][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][0][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][1][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][1][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][2][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][2][1]).split(":")[1]
                ))
            elif len(self.cc[self.keysList[q]]["Moves"]) == 2:
                print("{:<12} {:<10} {:<6} {:<6} {:<30} {:<45}".format(
                    self.keysList[q], self.cc[self.keysList[q]]["Type"], self.cc[self.keysList[q]]["HP"],
                    self.cc[self.keysList[q]]["Shiny"],
                    str(self.cc[self.keysList[q]]["Moves"][0][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][0][1]).split(":")[1],
                    str(self.cc[self.keysList[q]]["Moves"][1][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][1][1]).split(":")[1]
                ))
            elif len(self.cc[self.keysList[q]]["Moves"]) == 1:
                print("{:<12} {:<10} {:<6} {:<6} {:<30} ".format(
                    self.keysList[q], self.cc[self.keysList[q]]["Type"], self.cc[self.keysList[q]]["HP"],
                    self.cc[self.keysList[q]]["Shiny"],
                    str(self.cc[self.keysList[q]]["Moves"][0][0]) + ":" +
                    str(self.cc[self.keysList[q]]["Moves"][0][1]).split(":")[1],
                ))
            else:
                print("Your card doesn't have any moves.Please check your card")

    # Func viewAllShinyCards
    def viewAllShinyCards(self):
        """viewAllShinyCards prints the information of all the shiny cards in the deck by searching for cards which have
        shiny attribute equal to 1 """
        li = []  # Initialising an empty list
        for m in range(0, len(self.keysList)):
            isShiny = self.cc[self.keysList[m]]["Shiny"]
            if isShiny == 1:
                li.append(self.keysList[m])  # Adding shiny card to the list
            else:
                continue
        print("The following are the {} shiny cards in this deck:\n---------".format(len(li)))
        for card in li:
            print(
                "Card Name=> {} \nType=> {} \nHP=> {} \nShiny=> {} \nMoves=> {} \n---------------------------------"
                .format(card, self.cc[card]["Type"], self.cc[card]["HP"], self.cc[card]["Shiny"],
                        self.cc[card]["Moves"]))

    # Func viewAllByType returns the cards of a specific type
    def viewAllByType(self, theType):
        """viewAllByType prints the information of cards of a specific valid type by first checking if the type
        requested by user is valid or not"""
        typeList = ["Magi", "Water", "Fire", "Earth", "Air", "Astral"]
        if theType in typeList:  # if requested type is valid
            lis = []  # Initialising an empty list
            for v in range(0, len(self.keysList)):
                isType = self.cc[self.keysList[v]]["Type"]
                if isType == theType:
                    lis.append(self.keysList[v])  # Adding the card of given type to the list
                else:
                    continue
            if len(lis) == 0:  # if list is empty
                print("There is no card of type {} in this deck".format(theType))
            else:  # if list is not empty
                print("The following are the {} type cards in this deck:\n---------".format(theType))
                for card in lis:
                    print(
                        "Card Name=> {} \nType=> {} \nHP=> {} \nShiny=> {} \nMoves=> {} "
                        "\n-------------------------------------".format(
                            card, self.cc[card]["Type"], self.cc[card]["HP"], self.cc[
                                card]["Shiny"],
                            self.cc[card]["Moves"]))
        else:  # if requested type is not valid
            print("{} is not a valid type.Please try again.".format(theType))

    # Func getCards returns the cards in the deck as a collection
    def getCards(self):
        """getCards returns all cards held within the deck as a collection of type dictionary."""

        return "The following are the cards in this deck:\n{}".format(self.cc)

    # Func saveToFile saves the deck in the requested file
    def saveToFile(self, fileName):
        """saveToFile saves the Deck to the requested xlsx file that is called <fileName>.xlsx"""
        w = Workbook()
        sheet = w.active
        row = 1
        column = 1
        headerList = ["Name", 'Type', 'HP', 'Shiny', 'Move Name 1', 'Damage 1', 'Move Name 2', 'Damage 2',
                      'Move Name 3', 'Damage 3', 'Move Name 4', 'Damage 4', 'Move Name 5', 'Damage 5']
        for header in headerList:
            sheet.cell(row=row, column=column, value=header)
            column = column + 1
        row = 1
        for key, values in self.cc.items():
            sheet.cell(row=row + 1, column=1, value=key)
            column = 2
            for element in values.items():
                if element[0] != "Moves":
                    sheet.cell(row=row + 1, column=column, value=element[1])
                    column += 1
                else:
                    for move, damage in element[1]:
                        dString = damage.split(":")[1]
                        sheet.cell(row=row + 1, column=column, value=move)
                        column += 1
                        sheet.cell(row=row + 1, column=column, value=int(dString))
                        column += 1

            row += 1
        w.save(filename="/Users/sahibbhatia/Desktop/Programming_Fundamentals/{}.xlsx".format(fileName))
        # Change above path accordingly
        print("Deck has been saved to {}.xlsx".format(fileName))
